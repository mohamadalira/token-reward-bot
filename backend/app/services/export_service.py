import csv
import io
from typing import Any

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories import SponsorRepository


async def export_campaign_csv(session: AsyncSession, campaign_id: int, user_id: int) -> str:
    repo = SponsorRepository(session)
    sponsor = await repo.get_by_user_id(user_id)
    campaign = await repo.get_campaign(campaign_id)
    if not campaign or not sponsor or campaign.sponsor_id != sponsor.id:
        raise ValueError("Campaign not found")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["metric", "value"])
    writer.writerow(["title", campaign.channel_title])
    writer.writerow(["reward_per_join", campaign.reward_per_join])
    writer.writerow(["total_budget", campaign.total_budget])
    writer.writerow(["remaining_budget", campaign.remaining_budget])
    writer.writerow(["distributed_tokens", campaign.distributed_tokens])
    writer.writerow(["total_joins", campaign.total_joins])
    writer.writerow(["total_views", campaign.total_views])
    conv = (campaign.total_joins / campaign.total_views * 100) if campaign.total_views else 0
    writer.writerow(["conversion_rate", f"{conv:.2f}%"])
    return output.getvalue()


async def export_campaign_xlsx(session: AsyncSession, campaign_id: int, user_id: int) -> bytes:
    repo = SponsorRepository(session)
    campaign = await repo.get_campaign(campaign_id)
    sponsor = await repo.get_by_user_id(user_id)
    if not campaign or not sponsor or campaign.sponsor_id != sponsor.id:
        raise ValueError("Campaign not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Campaign Stats"
    rows = [
        ("عنوان", campaign.channel_title),
        ("پاداش هر عضویت", campaign.reward_per_join),
        ("بودجه کل", campaign.total_budget),
        ("باقیمانده", campaign.remaining_budget),
        ("توزیع شده", campaign.distributed_tokens),
        ("عضویت", campaign.total_joins),
        ("بازدید", campaign.total_views),
    ]
    for i, (k, v) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_campaign_pdf(session: AsyncSession, campaign_id: int, user_id: int) -> bytes:
    repo = SponsorRepository(session)
    campaign = await repo.get_campaign(campaign_id)
    sponsor = await repo.get_by_user_id(user_id)
    if not campaign or not sponsor or campaign.sponsor_id != sponsor.id:
        raise ValueError("Campaign not found")

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    y = 800
    c.setFont("Helvetica", 14)
    c.drawString(50, y, f"Campaign Report: {campaign.channel_title}")
    y -= 30
    c.setFont("Helvetica", 11)
    for label, val in [
        ("Reward/Join", campaign.reward_per_join),
        ("Budget", campaign.total_budget),
        ("Remaining", campaign.remaining_budget),
        ("Joins", campaign.total_joins),
        ("Views", campaign.total_views),
    ]:
        c.drawString(50, y, f"{label}: {val}")
        y -= 20
    c.save()
    return buf.getvalue()
